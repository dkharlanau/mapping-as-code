from __future__ import annotations

import csv
from pathlib import Path
from typing import Any, Iterable


def _truthy(value: Any) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "yes", "y", "x"}


def _rows_from_csv(path: Path) -> Iterable[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        yield from csv.DictReader(handle)


def _rows_from_xlsx(path: Path) -> Iterable[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Excel import requires: pip install 'mapping-as-code[excel]'") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows)]
    for row in rows:
        yield {headers[i]: row[i] for i in range(min(len(headers), len(row)))}


def import_tabular(path: str | Path, *, name: str, source_system: str, source_object: str, target_system: str, target_object: str) -> dict[str, Any]:
    input_path = Path(path)
    rows = _rows_from_xlsx(input_path) if input_path.suffix.lower() == ".xlsx" else _rows_from_csv(input_path)
    fields: list[dict[str, Any]] = []
    value_maps: dict[str, list[dict[str, str]]] = {}
    for index, row in enumerate(rows, start=1):
        source = str(row.get("source") or "").strip()
        target = str(row.get("target") or "").strip()
        if not source and not target:
            continue
        transform_type = str(row.get("transform") or "copy").strip().lower()
        reference = str(row.get("reference") or "").strip()
        field: dict[str, Any] = {
            "id": str(row.get("id") or f"field-{index:03d}").strip(),
            "source": {"path": source}, "target": {"path": target}, "transform": {"type": transform_type},
        }
        if reference:
            field["transform"]["reference"] = reference
            value_maps.setdefault(reference, [])
        if _truthy(row.get("required")):
            field["constraints"] = {"required": True}
        description = str(row.get("description") or "").strip()
        if description:
            field["description"] = description
        fields.append(field)
    return {
        "apiVersion": "mappingascode.dev/v1alpha1", "kind": "MappingSpec",
        "metadata": {"name": name, "version": "0.1.0"},
        "source": {"system": source_system, "object": source_object},
        "target": {"system": target_system, "object": target_object},
        "fields": fields, **({"valueMaps": value_maps} if value_maps else {}),
    }
