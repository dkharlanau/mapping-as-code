from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Any


class ImportErrorDetail(ValueError):
    """Raised when tabular mapping input is ambiguous or inconsistent."""


def _clean(value: Any) -> Any:
    if isinstance(value, str):
        value = value.strip()
        return value if value else None
    return value


def _as_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return value != 0
    return str(value).strip().lower() in {"1", "true", "yes", "y", "x", "required"}


def _slug(value: str) -> str:
    slug = re.sub(r"[^A-Za-z0-9_.-]+", "-", value.strip()).strip("-")
    return slug or "field"


def _normalized_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    for row in rows:
        cleaned = {str(key).strip(): _clean(value) for key, value in row.items() if key is not None}
        if any(value is not None for value in cleaned.values()):
            normalized.append(cleaned)
    return normalized


def _single(rows: list[dict[str, Any]], column: str) -> str:
    values = {str(row[column]) for row in rows if row.get(column) is not None}
    if not values:
        raise ImportErrorDetail(f"missing required workbook metadata column value: {column}")
    if len(values) > 1:
        raise ImportErrorDetail(f"inconsistent workbook metadata for {column}: {sorted(values)}")
    return values.pop()


def _read_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return _normalized_rows(list(csv.DictReader(handle)))


def _sheet_rows(sheet: Any) -> list[dict[str, Any]]:
    iterator = sheet.iter_rows(values_only=True)
    try:
        header_values = next(iterator)
    except StopIteration:
        return []
    headers = [str(value).strip() if value is not None else "" for value in header_values]
    rows: list[dict[str, Any]] = []
    for values in iterator:
        row = {headers[index]: value for index, value in enumerate(values) if index < len(headers) and headers[index]}
        rows.append(row)
    return _normalized_rows(rows)


def _read_xlsx(path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportErrorDetail(
            "XLSX import requires the optional dependency: pip install 'mapping-as-code[excel]'"
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    mapping_sheet = workbook["Mappings"] if "Mappings" in workbook.sheetnames else workbook.active
    mappings = _sheet_rows(mapping_sheet)
    value_maps = _sheet_rows(workbook["ValueMaps"]) if "ValueMaps" in workbook.sheetnames else []
    return mappings, value_maps


def read_tabular(
    path: str | Path, value_maps_path: str | Path | None = None
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    source = Path(path)
    if source.suffix.lower() in {".xlsx", ".xlsm"}:
        mappings, value_maps = _read_xlsx(source)
        if value_maps_path:
            value_maps = _read_csv(Path(value_maps_path))
        return mappings, value_maps
    if source.suffix.lower() != ".csv":
        raise ImportErrorDetail("supported tabular formats are .csv, .xlsx, and .xlsm")
    mappings = _read_csv(source)
    value_maps = _read_csv(Path(value_maps_path)) if value_maps_path else []
    return mappings, value_maps


def import_rows(
    rows: list[dict[str, Any]], value_map_rows: list[dict[str, Any]] | None = None
) -> dict[str, Any]:
    rows = _normalized_rows(rows)
    if not rows:
        raise ImportErrorDetail("mapping workbook contains no data rows")

    mapping_id = _single(rows, "mapping_id")
    source_system = _single(rows, "source_system")
    source_object = _single(rows, "source_object")
    target_system = _single(rows, "target_system")
    target_object = _single(rows, "target_object")

    fields: list[dict[str, Any]] = []
    required_targets: list[str] = []
    required_sources: list[str] = []

    for index, row in enumerate(rows, start=2):
        target_field = row.get("target_field")
        if target_field is None:
            raise ImportErrorDetail(f"row {index}: target_field is required")

        transform_type = str(row.get("transform") or "copy").strip().lower()
        source_field = row.get("source_field")
        if transform_type != "constant" and source_field is None:
            raise ImportErrorDetail(f"row {index}: source_field is required for {transform_type} transform")

        field_id = row.get("id")
        if field_id is None:
            left = str(source_field) if source_field is not None else "constant"
            field_id = _slug(f"{left}-to-{target_field}")

        field: dict[str, Any] = {
            "id": str(field_id),
            "target": {"field": str(target_field)},
            "transform": {"type": transform_type},
        }
        if source_field is not None:
            field["source"] = {"field": str(source_field)}

        if transform_type == "lookup" and row.get("reference") is not None:
            field["transform"]["reference"] = str(row["reference"])
        if transform_type == "constant":
            if row.get("value") is None:
                raise ImportErrorDetail(f"row {index}: constant transform requires value")
            field["transform"]["value"] = row["value"]
        if transform_type == "expression" and row.get("expression") is not None:
            field["transform"]["expression"] = str(row["expression"])

        rules: dict[str, Any] = {}
        if _as_bool(row.get("required_target")):
            required_targets.append(str(target_field))
            rules["required"] = True
        if _as_bool(row.get("required_source")) and source_field is not None:
            required_sources.append(str(source_field))
        if rules:
            field["rules"] = rules

        business = {
            key: row.get(key)
            for key in ("owner", "criticality", "rationale")
            if row.get(key) is not None
        }
        if business:
            field["business"] = business

        if _as_bool(row.get("allow_multiple_sources")):
            field["allow_multiple_sources"] = True
        fields.append(field)

    value_maps: dict[str, dict[Any, Any]] = {}
    for index, row in enumerate(_normalized_rows(value_map_rows or []), start=2):
        name = row.get("map")
        source_value = row.get("source")
        if name is None or source_value is None or "target" not in row or row.get("target") is None:
            raise ImportErrorDetail(f"value-map row {index}: map, source, and target are required")
        value_maps.setdefault(str(name), {})[source_value] = row["target"]

    document: dict[str, Any] = {
        "schema_version": "0.1",
        "mapping": {
            "id": mapping_id,
            "source": {
                "system": source_system,
                "object": source_object,
                "required_fields": list(dict.fromkeys(required_sources)),
            },
            "target": {
                "system": target_system,
                "object": target_object,
                "required_fields": list(dict.fromkeys(required_targets)),
            },
            "fields": fields,
        },
    }
    if value_maps:
        document["value_maps"] = value_maps
    return document


def import_tabular(path: str | Path, value_maps_path: str | Path | None = None) -> dict[str, Any]:
    rows, value_maps = read_tabular(path, value_maps_path=value_maps_path)
    return import_rows(rows, value_maps)
